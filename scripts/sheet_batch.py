#!/usr/bin/env python3
"""
sheet_batch.py -- shared spreadsheet-driven batch-run helper for the
per-.ply measurement scripts (fit_dab.py, dendro_tape.py, median_polygon.py,
median_polygon_2deg.py).

WHY THIS EXISTS
----------------
Each measurement script normally runs once per tree, with the tree's height
(if any) typed by hand into --slice-height and its .ply path typed by hand
as the positional argument -- fine for one-off work, but it doesn't batch,
and it ties every run to one specific file layout and column-naming scheme.
This module lets a script instead read tree ID / height / output column
straight from a spreadsheet (the CONFIG block at the top of each script) and
loop over every row in one run -- someone with different column names and
folders only has to edit that CONFIG block, not the measurement code.
That's the whole point: reusability for someone else's dataset.

Shared ACROSS the four Python measurement scripts -- this is plumbing, not
measurement logic, so it does NOT fall under the "two independent
implementations, deliberately kept separate" convention these scripts
otherwise follow. That convention is about Python vs R each independently
fitting the same tree as a correctness cross-check; sharing this file is
plain DRY, same reasoning as plot_style.R. NOT shared with the R side:
scripts/sheet_batch.R is an independently written R equivalent, so Python
and R measurement runs still don't share any code that could hide the same
bug in both.

WRITE-BACK SAFETY -- READ THIS BEFORE CHANGING write_back_all()
------------------------------------------------------------------
Uses openpyxl directly (load_workbook -> edit specific cells -> save), NOT
pandas.to_excel(). pandas.to_excel() REWRITES THE ENTIRE FILE from a bare
DataFrame and would silently destroy any other sheets/tabs, cell formatting,
comments, or data validation in a hand-maintained workbook -- unacceptable
for a private, irreplaceable working spreadsheet. openpyxl's load/save
round-trip only touches the specific cells this module writes to. All writes
for a run are batched into ONE load/edit/save cycle (see write_back_all),
not one per tree: faster, and it means a crashed run leaves the sheet either
fully updated or completely untouched -- never half-written.
"""
from __future__ import annotations
import os
from dataclasses import dataclass


@dataclass
class SheetConfig:
    """Bundles one script's CONFIG block (see that script's top-of-file
    CONFIG section for the actual values) into a single object to pass
    around. `height_col=None` means the script measures each .ply as an
    already-cut disc rather than cutting a band; `output_col=None` means
    write-back is skipped (results still go to --out CSV as usual)."""
    sheet_path: str
    tree_id_col: str
    ply_folder: str
    ply_filename_pattern: str = "{tree_id}.ply"    # {tree_id} required; {site} optional
    site_label: str | None = None
    height_col: str | None = None
    output_col: str | None = None
    sheet_name: str | int = 0


def _require_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        import sys
        sys.exit("Missing dependency: pip install openpyxl  (see requirements.txt)")


def _col_letter(ws, header_name: str) -> str:
    """Find the column letter whose row-1 header cell equals `header_name`."""
    from openpyxl.utils import get_column_letter
    for cell in ws[1]:
        if cell.value == header_name:
            return get_column_letter(cell.column)
    raise KeyError(f"Column '{header_name}' not found in the header row of {ws.title!r}.")


def load_manifest(cfg: SheetConfig) -> list[dict]:
    """Read (tree_id, height, ply path, sheet row) for every row that has a
    tree ID and, if cfg.height_col is set, a non-empty height. Rows missing
    either are skipped with a printed reason -- normal mid-project state
    (not every tree has been picked/cut yet), not necessarily an error."""
    _require_openpyxl()
    import openpyxl

    wb = openpyxl.load_workbook(cfg.sheet_path, data_only=True, read_only=True)
    ws = wb[cfg.sheet_name] if isinstance(cfg.sheet_name, str) else wb.worksheets[cfg.sheet_name]

    id_col = _col_letter(ws, cfg.tree_id_col)
    height_col = _col_letter(ws, cfg.height_col) if cfg.height_col else None

    manifest = []
    for r in range(2, ws.max_row + 1):
        tree_id = ws[f"{id_col}{r}"].value
        if tree_id is None or str(tree_id).strip() == "":
            continue
        height = ws[f"{height_col}{r}"].value if height_col else None
        if cfg.height_col and (height is None or height == ""):
            print(f"[skip] {tree_id}: no value in '{cfg.height_col}' yet")
            continue
        fname = cfg.ply_filename_pattern.format(tree_id=tree_id, site=cfg.site_label or "")
        path = os.path.join(cfg.ply_folder, fname)
        if not os.path.exists(path):
            print(f"[skip] {tree_id}: {path} not found")
            continue
        manifest.append({
            "tree_id": str(tree_id),
            "height": float(height) if height is not None else None,
            "path": path,
            "row": r,
        })
    wb.close()
    return manifest


def write_back_all(cfg: SheetConfig, updates: list[tuple[int, float]]) -> None:
    """Write every (sheet_row, value) pair into cfg.output_col in ONE
    load/edit/save cycle. No-op if output_col isn't configured or there's
    nothing to write -- callers can call this unconditionally."""
    if not cfg.output_col or not updates:
        return
    _require_openpyxl()
    import openpyxl

    wb = openpyxl.load_workbook(cfg.sheet_path)   # data_only=False: don't disturb formulas elsewhere
    ws = wb[cfg.sheet_name] if isinstance(cfg.sheet_name, str) else wb.worksheets[cfg.sheet_name]
    out_col = _col_letter(ws, cfg.output_col)
    for row_number, value in updates:
        ws[f"{out_col}{row_number}"] = value
    wb.save(cfg.sheet_path)
    wb.close()
    print(f"Wrote {len(updates)} value(s) into '{cfg.output_col}' -> {cfg.sheet_path}")
